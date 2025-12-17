import pandas as pd
import math
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import tempfile
import os, json
from io import BytesIO
from openai import OpenAI
import hashlib
import time
from typing import Any, Dict, List , Optional
client = OpenAI(api_key="")

IMPORTANT_SHEETS = [
    "РУП",
    "Нагруженность преподователей",
    "Группы и направления",
    "Аудитории",
    "Правила составления",
]

MAX_ROWS_PER_SHEET = 80

SHEET_COLUMNS = {
    "Нагруженность преподователей": ["Индекс", "Дисциплина", "ФИО преподавателя", "группа", "семестр", "количество часов"],
    "Правила составления": ["Параметр", "Описание", "Пример", "Код", "Правило", "Вес штрафа"],
    "Группы и направления": ["Группа", "Семестр", "Размер группы"],
    "Аудитории": ["Аудитория", "Назначение", "Вместимость"],
    # РУП зависит от вашей структуры — оставьте основные
    "РУП": None,  # None => берем все колонки, но можно сузить позже
}
_AI_CACHE: Dict[str, Dict[str, Any]] = {}
_AI_CACHE_TTL_SEC = 60 * 60  # 1 час


def split_load_by_semester(df_teachers: pd.DataFrame) -> Dict[int, pd.DataFrame]:
    if "семестр" not in df_teachers.columns:
        return {}

    result = {}
    for sem in sorted(df_teachers["семестр"].dropna().unique()):
        try:
            sem_int = int(sem)
        except:
            continue
        result[sem_int] = df_teachers[df_teachers["семестр"] == sem].copy()
    return result
def sheet_preview(df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
    df2 = df.copy()

    wanted = SHEET_COLUMNS.get(sheet_name)
    if wanted:
        existing = [c for c in wanted if c in df2.columns]
        df2 = df2[existing]

    df2 = df2.head(MAX_ROWS_PER_SHEET).copy()
    df2.columns = [str(c) for c in df2.columns]
    rows = df2.to_dict(orient="records")

    # excel_row: +2 потому что заголовок в строке 1, первая запись обычно со 2 строки
    rows = [{"excel_row": i + 2, **r} for i, r in enumerate(rows)]

    return {
        "columns": [str(c) for c in df2.columns],
        "row_count": int(len(df)),
        "rows_preview": rows,
    }

def build_payload(excel_bytes: bytes) -> Dict[str, Any]:
    xls = pd.ExcelFile(BytesIO(excel_bytes))
    payload = {
        "sheet_names": xls.sheet_names,
        "sheets": {},
        "meta": {
            "max_rows_per_sheet": MAX_ROWS_PER_SHEET,
            "columns_policy": SHEET_COLUMNS,
        }
    }

    for name in IMPORTANT_SHEETS:
        if name not in xls.sheet_names:
            payload["sheets"][name] = {"error": "sheet_missing"}
            continue
        df = pd.read_excel(xls, sheet_name=name)
        payload["sheets"][name] = sheet_preview(df, name)

    return payload

REPORT_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "summary": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "errors": {"type": "integer"},
                "warnings": {"type": "integer"},
                "notes": {"type": "integer"},
            },
            "required": ["errors", "warnings", "notes"],
        },
        "errors": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "sheet": {"type": "string"},
                    "excel_row": {"type": ["integer", "null"]},
                    "column": {"type": ["string", "null"]},
                    "code": {"type": "string"},
                    "message": {"type": "string"},
                    "evidence": {"type": "string"},
                },
                "required": ["sheet", "excel_row", "column", "code", "message", "evidence"],
            },
        },
        "warnings": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "sheet": {"type": ["string", "null"]},
                    "excel_row": {"type": ["integer", "null"]},
                    "column": {"type": ["string", "null"]},
                    "code": {"type": "string"},
                    "message": {"type": "string"},
                    "evidence": {"type": "string"},
                },
                "required": ["sheet", "excel_row", "column", "code", "message", "evidence"],
            },
        },
        "notes": {"type": "array", "items": {"type": "string"}},
        "rules_feedback": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "params": {"type": "array", "items": {"type": "string"}},
                "hard": {"type": "array", "items": {"type": "string"}},
                "soft": {"type": "array", "items": {"type": "string"}},
                "issues": {"type": "array", "items": {"type": "string"}},
                "suggestions": {"type": "array", "items": {"type": "string"}},
            },
            "required": ["params", "hard", "soft", "issues", "suggestions"],
        },
    },
    "required": ["summary", "errors", "warnings", "notes", "rules_feedback"],
}

def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()

REQUIRED_RULE_PARAMS = [
    "Direction_Type","Study_Days_Per_Week","Max_Lessons_Per_Day",
    "Min_Lessons_Per_Day","Lesson_Duration_Min","Semester_Weeks","Shift_Type",
]
def _return_json_file(payload: dict, filename: str):
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".json")
    with open(tmp.name, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    return FileResponse(
        path=tmp.name,
        media_type="application/json",
        filename=filename
    )
def _norm_str(x) -> str:
    return "" if x is None or (isinstance(x, float) and pd.isna(x)) else str(x).strip()

def _to_int(x) -> Optional[int]:
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return None
        return int(float(x))
    except:
        return None

def _to_float(x) -> Optional[float]:
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return None
        return float(x)
    except:
        return None

def _rules_map(df_rules: pd.DataFrame) -> Dict[str, str]:
    """
    Ожидаем лист 'Правила составления' в формате:
    Параметр | Описание | Пример | ...
    Пример используем как значение.
    """
    mp: Dict[str, str] = {}
    if df_rules is None or df_rules.empty:
        return mp
    if "Параметр" not in df_rules.columns:
        return mp
    ex_col = "Пример" if "Пример" in df_rules.columns else None

    for _, r in df_rules.iterrows():
        p = _norm_str(r.get("Параметр"))
        if not p:
            continue
        ex = _norm_str(r.get(ex_col)) if ex_col else ""
        mp[p] = ex
    return mp

def logic_precheck_full(excel_bytes: bytes) -> List[Dict[str, Any]]:
    xls = pd.ExcelFile(BytesIO(excel_bytes))

    df_rup   = pd.read_excel(xls, sheet_name="РУП")
    df_load  = pd.read_excel(xls, sheet_name="Нагруженность преподователей")
    df_groups= pd.read_excel(xls, sheet_name="Группы и направления")
    df_rooms = pd.read_excel(xls, sheet_name="Аудитории")
    df_rules = pd.read_excel(xls, sheet_name="Правила составления")

    errors: List[Dict[str, Any]] = []

    # ----------------------------
    # 0) Настройки из "Правила составления"
    # ----------------------------
    rules = _rules_map(df_rules)

    WEEKS = _to_int(rules.get("Semester_Weeks")) or 16
    DAYS_PER_WEEK = _to_int(rules.get("Study_Days_Per_Week")) or 5
    MAX_LESSONS_PER_DAY = _to_int(rules.get("Max_Lessons_Per_Day")) or 5
    LESSON_MIN = _to_int(rules.get("Lesson_Duration_Min")) or 90

    # Переводим "кол-во часов" в пары.
    # Предположение: "количество часов" в Excel = АСТРОНОМИЧЕСКИЕ минуты? Нет.
    # Обычно в РУП/нагрузке "часы" = академические (45 мин) или астрономические (60 мин).
    # Выберите режим ниже одним флагом:
    HOURS_ARE_ACADEMIC_45 = True  # <-- если у вас "час" = 45 минут (часто так в колледжах)

    MINUTES_PER_HOUR = 45 if HOURS_ARE_ACADEMIC_45 else 60
    PAIR_MIN = LESSON_MIN
    SLOTS_PER_WEEK = DAYS_PER_WEEK * MAX_LESSONS_PER_DAY

    # ----------------------------
    # 1) Карта групп
    # ----------------------------
    if "Группа" not in df_groups.columns:
        return [{
            "sheet": "Группы и направления",
            "excel_row": None,
            "column": "Группа",
            "code": "MISSING_COLUMN",
            "message": "На листе 'Группы и направления' нет колонки 'Группа'.",
            "evidence": ""
        }]

    # Текущий семестр группы (важно!)
    group_current_sem: Dict[str, Optional[int]] = {}
    group_size: Dict[str, int] = {}

    for i, r in df_groups.iterrows():
        g = _norm_str(r.get("Группа"))
        if not g:
            continue
        cur_sem = _to_int(r.get("Семестр")) if "Семестр" in df_groups.columns else None
        size = _to_int(r.get("Размер группы")) if "Размер группы" in df_groups.columns else None
        group_current_sem[g] = cur_sem
        group_size[g] = size if size is not None else 25

    # ----------------------------
    # 2) Капацитеты аудиторий: проверка что группе вообще есть куда сесть
    # ----------------------------
    room_caps: List[int] = []
    if not df_rooms.empty and "Вместимость" in df_rooms.columns:
        for _, r in df_rooms.iterrows():
            cap = _to_int(r.get("Вместимость"))
            if cap is not None:
                room_caps.append(cap)
    max_cap = max(room_caps) if room_caps else 0

    for g, size in group_size.items():
        if max_cap > 0 and size > max_cap:
            errors.append({
                "sheet": "Группы и направления",
                "excel_row": None,
                "column": "Размер группы",
                "code": "NO_ROOM_FOR_GROUP",
                "message": f"Для группы '{g}' размер={size}, но максимальная вместимость аудитории={max_cap}.",
                "evidence": f"группа={g}, размер={size}, max_room_capacity={max_cap}"
            })

    # ----------------------------
    # 3) Проверки нагрузки + расчёт недельной загрузки
    # ----------------------------
    required_cols = ["группа", "семестр", "количество часов", "ФИО преподавателя", "Дисциплина"]
    for c in required_cols:
        if c not in df_load.columns:
            errors.append({
                "sheet": "Нагруженность преподователей",
                "excel_row": None,
                "column": c,
                "code": "MISSING_COLUMN",
                "message": f"На листе 'Нагруженность преподователей' нет колонки '{c}'.",
                "evidence": ""
            })
    if errors:
        return errors

    weekly_pairs_by_group: Dict[str, int] = {}
    weekly_pairs_by_teacher: Dict[str, int] = {}

    for i, r in df_load.iterrows():
        excel_row = i + 2

        grp = _norm_str(r.get("группа"))
        sem = _to_int(r.get("семестр"))
        teacher = _norm_str(r.get("ФИО преподавателя"))
        subj = _norm_str(r.get("Дисциплина"))
        hours = _to_float(r.get("количество часов"))

        # 3.1 Группа существует?
        if grp and grp not in group_current_sem:
            errors.append({
                "sheet": "Нагруженность преподователей",
                "excel_row": excel_row,
                "column": "группа",
                "code": "UNKNOWN_GROUP",
                "message": f"Группа '{grp}' отсутствует на листе 'Группы и направления'.",
                "evidence": f"группа={grp}"
            })
            continue

        # 3.2 Семестр дисциплины должен быть числом >=1
        if sem is None or sem < 1:
            errors.append({
                "sheet": "Нагруженность преподователей",
                "excel_row": excel_row,
                "column": "семестр",
                "code": "SEMESTER_INVALID",
                "message": f"Некорректный семестр дисциплины: '{r.get('семестр')}'. Должно быть число >= 1.",
                "evidence": f"группа={grp}, семестр={r.get('семестр')}"
            })
            continue

        # 3.3 Часы должны быть > 0
        if hours is None or hours <= 0:
            errors.append({
                "sheet": "Нагруженность преподователей",
                "excel_row": excel_row,
                "column": "количество часов",
                "code": "HOURS_INVALID",
                "message": f"Некорректное количество часов: '{r.get('количество часов')}'. Должно быть > 0.",
                "evidence": f"группа={grp}, дисциплина={subj}, часы={r.get('количество часов')}"
            })
            continue

        # 3.4 Логика: недельную нагрузку считаем ТОЛЬКО для текущего семестра группы
        cur_sem = group_current_sem.get(grp)
        if cur_sem is None:
            # Если у группы не указан текущий семестр — не считаем weekly overload, но и не считаем ошибкой.
            continue

        if sem != int(cur_sem):
            # Это нормальная ситуация: нагрузка расписана по всем семестрам.
            continue

        # 3.5 Перевод часов -> пары семестра -> пары в неделю
        total_minutes = hours * MINUTES_PER_HOUR
        total_pairs_sem = math.ceil(total_minutes / PAIR_MIN)
        pairs_per_week = math.ceil(total_pairs_sem / WEEKS)

        weekly_pairs_by_group[grp] = weekly_pairs_by_group.get(grp, 0) + pairs_per_week
        if teacher:
            weekly_pairs_by_teacher[teacher] = weekly_pairs_by_teacher.get(teacher, 0) + pairs_per_week

    # ----------------------------
    # 4) Проверка: "физически не влезает по слотам" (группа)
    # ----------------------------
    for grp, pairs_w in weekly_pairs_by_group.items():
        if pairs_w > SLOTS_PER_WEEK:
            errors.append({
                "sheet": "Нагруженность преподователей",
                "excel_row": None,
                "column": None,
                "code": "GROUP_OVERLOAD_WEEKLY",
                "message": (
                    f"Группа '{grp}' (текущий семестр={group_current_sem.get(grp)}) требует "
                    f"≈{pairs_w} пар/нед, но максимум слотов {SLOTS_PER_WEEK} "
                    f"({DAYS_PER_WEEK} дней × {MAX_LESSONS_PER_DAY} пар). Расписание невозможно."
                ),
                "evidence": f"группа={grp}, pairs_per_week≈{pairs_w}, max={SLOTS_PER_WEEK}"
            })

    # ----------------------------
    # 5) (Опционально) Проверка: перегрузка преподавателя по слотам (упрощённо)
    # ----------------------------
    # Это грубая логика: у преподавателя максимум те же SLOTS_PER_WEEK (если он работает 5x5).
    # Если у вас есть отдельные нормы/ставки — вынесите в правила.
    for t, pairs_w in weekly_pairs_by_teacher.items():
        if pairs_w > SLOTS_PER_WEEK:
            errors.append({
                "sheet": "Нагруженность преподователей",
                "excel_row": None,
                "column": "ФИО преподавателя",
                "code": "TEACHER_OVERLOAD_WEEKLY",
                "message": (
                    f"Преподаватель '{t}' в текущих семестрах групп требует ≈{pairs_w} пар/нед, "
                    f"что больше максимума {SLOTS_PER_WEEK}. Проверьте нагрузку/ставки."
                ),
                "evidence": f"teacher={t}, pairs_per_week≈{pairs_w}, max={SLOTS_PER_WEEK}"
            })

    return errors
def local_precheck(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    errors = []

    # 1) обязательные листы
    for sh in IMPORTANT_SHEETS:
        if payload["sheets"].get(sh, {}).get("error") == "sheet_missing":
            errors.append({
                "sheet": sh, "excel_row": None, "column": None,
                "code": "SHEET_MISSING",
                "message": f"Отсутствует обязательный лист: {sh}",
                "evidence": ""
            })

    # 2) обязательные параметры в правилах
    rules = payload["sheets"].get("Правила составления", {})
    if "rows_preview" in rules:
        param_to_example = {}
        for r in rules["rows_preview"]:
            p = str(r.get("Параметр") or "").strip()
            ex = "" if r.get("Пример") is None else str(r.get("Пример")).strip()
            if p:
                param_to_example[p] = ex

        for p in REQUIRED_RULE_PARAMS:
            if p not in param_to_example:
                errors.append({
                    "sheet": "Правила составления", "excel_row": None, "column": "Параметр",
                    "code": "MISSING_RULE_PARAM",
                    "message": f"Отсутствует обязательный параметр: {p}",
                    "evidence": ""
                })
            elif param_to_example[p] == "":
                errors.append({
                    "sheet": "Правила составления", "excel_row": None, "column": "Пример",
                    "code": "EMPTY_RULE_EXAMPLE",
                    "message": f"Параметр '{p}' есть, но 'Пример' пуст.",
                    "evidence": f"{p}: Пример пуст"
                })
    return errors
SYSTEM_PROMPT = """
Ты — эксперт по учебным планам колледжа и автоматической генерации расписаний.
Твоя задача — СТРОГО проверить Excel-файл перед генерацией расписания.

Тебе передаётся JSON-превью Excel-листов (РУП, Нагруженность преподователей,
Группы и направления, Аудитории, Правила составления).

Работай как формальный валидатор данных.
Не выдумывай данные. Не додумывай значения.
Проверяй ТОЛЬКО то, что реально есть в Excel.

Правило: любое обязательное поле null/NaN/""/пробелы => ERROR.
Если есть хотя бы 1 error => генерацию расписания запретить.

Обязательные листы:
- РУП
- Нагруженность преподователей
- Группы и направления
- Аудитории
- Правила составления
Отсутствие листа => критическая ошибка.

Лист "Правила составления":
обязательные параметры: Direction_Type, Study_Days_Per_Week, Max_Lessons_Per_Day,
Min_Lessons_Per_Day, Lesson_Duration_Min, Semester_Weeks, Shift_Type.
Если у параметра отсутствует/пуст "Пример" => ERROR.
Проверяй значения (диапазоны и Min<=Max).
H* должны быть HARD, S* должны быть SOFT/или пустой вес допускается как SOFT.

Верни результат СТРОГО по JSON schema. Никакого текста вне JSON.
"""


def ai_validate_excel(excel_bytes: bytes) -> Dict[str, Any]:
    file_hash = _sha256(excel_bytes)

    # кэш (если один и тот же файл гоняете несколько раз)
    cached = _AI_CACHE.get(file_hash)
    if cached and (time.time() - cached["ts"] < _AI_CACHE_TTL_SEC):
        return cached["report"]

    payload = build_payload(excel_bytes)

    # локальные ошибки — сразу возвращаем без оплаты
    local_errors = local_precheck(payload)
    if local_errors:
        report = {
            "summary": {"errors": len(local_errors), "warnings": 0, "notes": 0},
            "errors": local_errors,
            "warnings": [],
            "notes": [],
            "rules_feedback": {"params": [], "hard": [], "soft": [], "issues": [], "suggestions": []},
        }
        _AI_CACHE[file_hash] = {"ts": time.time(), "report": report}
        return report

    resp = client.responses.create(
        model="gpt-4.1-mini",
        input=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False)}
        ],
        temperature=0,
        max_output_tokens=900,  # держим низко => дешевле
        text={
            "format": {
                "type": "json_schema",
                "name": "validation_report",     # <-- это и лечит вашу ошибку
                "schema": REPORT_SCHEMA
            }
        }
    )

    report = json.loads(resp.output_text)
    _AI_CACHE[file_hash] = {"ts": time.time(), "report": report}
    return report


class ScheduleOptimizer:
    def __init__(self, rup_df, teachers_df, groups_df, rooms_df, rules_df):
        self.rup = rup_df
        self.teachers = teachers_df
        self.groups = groups_df
        self.rooms = rooms_df
        
        # Константы
        self.WEEKS = 16
        self.DAYS_PER_WEEK = 5
        self.MAX_PAIRS = 5
        self.PAIR_DURATION = 1.5 
        
        # Кэш размеров групп
        self.group_sizes = {}
        self._cache_group_sizes()

    def _cache_group_sizes(self):
        if not self.groups.empty and 'Группа' in self.groups.columns:
            for _, row in self.groups.iterrows():
                g_name = str(row['Группа']).strip()
                try: self.group_sizes[g_name] = int(row['Размер группы'])
                except: self.group_sizes[g_name] = 25

    def get_group_size(self, group_name):
        return self.group_sizes.get(group_name, 25)

    def calculate_weekly_needs(self, week_num):
        needs = []
        
        for _, row in self.teachers.iterrows():
            if pd.isna(row['Дисциплина']): continue
            subject = str(row['Дисциплина']).strip()
            teacher = str(row['ФИО преподавателя']).strip()
            group = str(row['группа']).strip()
            
            if not subject or not teacher or not group:
                continue
            
            try: total_hours = float(row['количество часов'])
            except: continue
            if total_hours == 0: continue

            # Расчет пар
            total_pairs_sem = math.ceil(total_hours / self.PAIR_DURATION)
            
            # Равномерное распределение
            cumulative_pairs_current = math.ceil((total_pairs_sem * week_num) / self.WEEKS)
            cumulative_pairs_prev = math.ceil((total_pairs_sem * (week_num - 1)) / self.WEEKS)
            pairs_this_week = cumulative_pairs_current - cumulative_pairs_prev
            
            if pairs_this_week == 0: continue

            subj_lower = subject.lower()
            is_sport = any(x in subj_lower for x in ['физическ', 'физк', 'спорт', 'нвп'])

            needs.append({
                'group': group,
                'subject': subject,
                'teacher': teacher,
                'pairs_count': pairs_this_week,
                'is_sport': is_sport,
                'id': f"{subject}_{teacher}_{is_sport}"
            })
        return needs

    def get_suitable_room(self, required_capacity, is_sport, occupied_rooms, tolerance=0):
        """
        tolerance: сколько человек может "стоять", если не влезают сидя.
        """
        candidates = []
        for _, room in self.rooms.iterrows():
            r_name = str(room['Аудитория']).strip()
            if r_name in occupied_rooms: continue

            try: r_capacity = int(room['Вместимость'])
            except: r_capacity = 30
            
            r_type = str(room['Назначение']) if pd.notna(room['Назначение']) else "Общая"
            is_gym_room = 'спорт' in r_name.lower() or 'физра' in str(r_type).lower()

            if is_sport:
                if is_gym_room: return r_name 
                continue 
            
            if is_gym_room: continue 
            
            # Проверка вместимости с учетом допуска (tolerance)
            if (r_capacity + tolerance) >= required_capacity:
                waste = r_capacity - required_capacity
                candidates.append((r_name, waste))
        
        # Сортируем: сначала те, где меньше всего пустых мест, но они положительные (waste может быть отрицательным при tolerance)
        candidates.sort(key=lambda x: x[1])
        return candidates[0][0] if candidates else None

    def _group_into_flows(self, needs):
        grouped = {}
        for item in needs:
            for _ in range(item['pairs_count']):
                key = item['id']
                if key not in grouped: grouped[key] = []
                grouped[key].append(item)

        tasks = []
        for key, items in grouped.items():
            items.sort(key=lambda x: x['group'])
            i = 0
            while i < len(items):
                current = items[i]
                next_item = items[i+1] if (i + 1) < len(items) else None
                
                if next_item and current['group'] != next_item['group']:
                    task = {
                        'groups': [current['group'], next_item['group']],
                        'subject': current['subject'],
                        'teacher': current['teacher'],
                        'is_sport': current['is_sport'],
                        'is_flow': True
                    }
                    tasks.append(task)
                    i += 2 
                else:
                    task = {
                        'groups': [current['group']],
                        'subject': current['subject'],
                        'teacher': current['teacher'],
                        'is_sport': current['is_sport'],
                        'is_flow': False
                    }
                    tasks.append(task)
                    i += 1
        
        # Сортировка: Потоки - первыми
        tasks.sort(key=lambda x: (not x['is_flow'], x['is_sport']), reverse=False)
        return tasks

    def calculate_slot_score(self, day, pair, groups, schedule_map):
        PENALTY_WINDOW = 100
        PENALTY_EDGE = 5
        BONUS_ADJACENT = -50
        
        score = 0
        if pair == 1 or pair == self.MAX_PAIRS: score += PENALTY_EDGE

        has_any_adjacent = False
        creates_window = False

        for grp in groups:
            grp_sched = schedule_map[day].get(grp, {})
            if not grp_sched:
                if pair in [2, 3]: score -= 10
                continue
            
            is_adjacent = (pair - 1 in grp_sched) or (pair + 1 in grp_sched)
            if is_adjacent: has_any_adjacent = True
            
            occupied_pairs = list(grp_sched.keys()) + [pair]
            min_p, max_p = min(occupied_pairs), max(occupied_pairs)
            span = max_p - min_p + 1
            if span > len(occupied_pairs):
                creates_window = True
        
        if creates_window: score += PENALTY_WINDOW
        if has_any_adjacent: score += BONUS_ADJACENT
            
        return score

    def _place_single_task(self, task_groups, subject, teacher, is_sport, schedule, teacher_busy, room_busy, group_schedule_map):
        """
        Двухэтапная попытка размещения
        """
        total_students = sum([self.get_group_size(g) for g in task_groups])
        
        # --- ПРОХОД 1: "Красивый" (Строгие правила) ---
        best_slot = None
        min_score = float('inf')

        for day in range(1, self.DAYS_PER_WEEK + 1):
            # Строгий лимит пар (макс 4)
            if len(group_schedule_map[day].get(task_groups[0], {})) >= 4: continue

            for pair in range(1, self.MAX_PAIRS + 1):
                if teacher in teacher_busy[day][pair]: continue
                
                groups_busy = False
                for g in task_groups:
                    if g in schedule[day][pair]: groups_busy = True; break
                if groups_busy: continue
                
                # Строгая вместимость (tolerance=0)
                suitable_room = self.get_suitable_room(total_students, is_sport, room_busy[day][pair], tolerance=0)
                if not suitable_room: continue

                score = self.calculate_slot_score(day, pair, task_groups, group_schedule_map)
                if score < min_score:
                    min_score = score
                    best_slot = (day, pair, suitable_room)
        
        if best_slot:
            self._commit_slot(best_slot, task_groups, subject, teacher, schedule, teacher_busy, room_busy, group_schedule_map)
            return True, "OK"

        # --- ПРОХОД 2: "Силовой" (Desperate Mode) ---
        # Если не вышло красиво, разрешаем:
        # 1. 5 пар в день
        # 2. Переполнение аудитории на 8 человек
        # 3. Любое окно (игнорируем score)
        
        for day in range(1, self.DAYS_PER_WEEK + 1):
            # Relaxed limit: разрешаем 5 пар, если очень надо
            if len(group_schedule_map[day].get(task_groups[0], {})) >= 5: continue

            for pair in range(1, self.MAX_PAIRS + 1):
                if teacher in teacher_busy[day][pair]: continue
                
                groups_busy = False
                for g in task_groups:
                    if g in schedule[day][pair]: groups_busy = True; break
                if groups_busy: continue
                
                # RELAXED вместимость (tolerance=8)
                suitable_room = self.get_suitable_room(total_students, is_sport, room_busy[day][pair], tolerance=8)
                
                if suitable_room:
                    # Сразу берем первое попавшееся (Greedy)
                    best_slot = (day, pair, suitable_room)
                    self._commit_slot(best_slot, task_groups, subject, teacher, schedule, teacher_busy, room_busy, group_schedule_map)
                    return True, "Forced"

        return False, "No Room/Time"

    def _commit_slot(self, slot, groups, subject, teacher, schedule, teacher_busy, room_busy, group_schedule_map):
        day, pair, room = slot
        for g in groups:
            schedule[day][pair][g] = {
                'subject': subject,
                'teacher': teacher,
                'room': room,
                'is_flow': len(groups) > 1
            }
            if g not in group_schedule_map[day]: group_schedule_map[day][g] = {}
            group_schedule_map[day][g][pair] = True

        teacher_busy[day][pair].add(teacher)
        room_busy[day][pair].add(room)

    def generate_week_schedule(self, week_num):
        raw_needs = self.calculate_weekly_needs(week_num)
        tasks = self._group_into_flows(raw_needs)
        
        schedule = {d: {p: {} for p in range(1, self.MAX_PAIRS + 1)} for d in range(1, self.DAYS_PER_WEEK + 1)}
        group_schedule_map = {d: {} for d in range(1, self.DAYS_PER_WEEK + 1)}
        teacher_busy = {d: {p: set() for p in range(1, self.MAX_PAIRS + 1)} for d in range(1, self.DAYS_PER_WEEK + 1)}
        room_busy = {d: {p: set() for p in range(1, self.MAX_PAIRS + 1)} for d in range(1, self.DAYS_PER_WEEK + 1)}
        
        unscheduled = []

        for task in tasks:
            groups = task['groups']
            subject = task['subject']
            teacher = task['teacher']
            is_sport = task['is_sport']
            
            # 1. Пробуем поставить задачу (поток или соло)
            success, msg = self._place_single_task(groups, subject, teacher, is_sport, schedule, teacher_busy, room_busy, group_schedule_map)

            # 2. Если это был ПОТОК и не вышло -> Разбиваем
            if not success and task['is_flow']:
                # print(f"DEBUG: Разбиваем поток {groups} ({msg})")
                split_failed_groups = []
                for single_group in groups:
                    sub_success, sub_msg = self._place_single_task([single_group], subject, teacher, is_sport, schedule, teacher_busy, room_busy, group_schedule_map)
                    if not sub_success:
                        split_failed_groups.append(single_group)
                
                if split_failed_groups:
                     unscheduled.append(f"Неделя {week_num} | {', '.join(split_failed_groups)}: {subject} (ERR: {sub_msg})")
            
            elif not success:
                unscheduled.append(f"Неделя {week_num} | {', '.join(groups)}: {subject} ({teacher}) (ERR: {msg})")

        return schedule, unscheduled

    def generate_semester(self):
        semester_schedule = {}
        all_errors = []
        print(f"INFO: Старт генерации семестра ({self.WEEKS} недель)...")
        for w in range(1, self.WEEKS + 1):
            sch, errs = self.generate_week_schedule(w)
            semester_schedule[w] = sch
            all_errors.extend(errs)
        return semester_schedule, all_errors

    def save_semester_to_excel(self, semester_schedule, output_filename="Расписание_Семестр.xlsx"):
        wb = Workbook()
        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        flow_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") 
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        days_names = {1: "ПОНЕДЕЛЬНИК", 2: "ВТОРНИК", 3: "СРЕДА", 4: "ЧЕТВЕРГ", 5: "ПЯТНИЦА"}

        all_groups = set()
        for w in semester_schedule:
            for d in semester_schedule[w]:
                for p in semester_schedule[w][d]:
                    for g in semester_schedule[w][d][p]: all_groups.add(g)
        sorted_groups = sorted(list(all_groups))

        for w in range(1, self.WEEKS + 1):
            if w not in semester_schedule: continue
            ws = wb.create_sheet(f"Неделя {w}")
            schedule = semester_schedule[w]
            headers = ["День", "Пара"] + sorted_groups
            ws.append(headers)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font; cell.fill = header_fill
                cell.alignment = center_align; cell.border = thin_border
                if col_num <= 2: ws.column_dimensions[get_column_letter(col_num)].width = 5
                else: ws.column_dimensions[get_column_letter(col_num)].width = 20

            current_row = 2
            for day in range(1, self.DAYS_PER_WEEK + 1):
                start_row = current_row
                for pair in range(1, self.MAX_PAIRS + 1):
                    ws.cell(row=current_row, column=1, value=days_names.get(day)).border = thin_border
                    ws.cell(row=current_row, column=2, value=pair).alignment = center_align
                    ws.cell(row=current_row, column=2).border = thin_border
                    for i, group in enumerate(sorted_groups):
                        col_idx = 3 + i
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.border = thin_border; cell.alignment = center_align
                        if group in schedule[day][pair]:
                            info = schedule[day][pair][group]
                            cell.value = f"{info['subject']}\n({info['teacher']})\n{info['room']}"
                            if info.get('is_flow'): cell.fill = flow_fill 
                        else: cell.value = ""
                    current_row += 1
                ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row-1, end_column=1)
                day_cell = ws.cell(row=start_row, column=1)
                day_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
                day_cell.font = Font(bold=True)
        try:
            wb.save(output_filename)
            print(f"\nФайл успешно сохранен: {output_filename}")
        except Exception as e: print(f"\nОШИБКА сохранения: {e}")
    
    def save_semester_to_json(self, semester_schedule, output_json_path):
        """
        Сохраняет semester_schedule в JSON под фронт Vue:
        {
          "groups": [...],
          "teachers": [...],
          "weeks": [
            {
              "week_number": 1,
              "days": [
                {"day_name": "ПОНЕДЕЛЬНИК", "pairs": [ { "pair":1, "slots": {...} }, ... ]},
                ...
              ]
            },
            ...
          ]
        }
        """
        days_names = {1: "ПОНЕДЕЛЬНИК", 2: "ВТОРНИК", 3: "СРЕДА", 4: "ЧЕТВЕРГ", 5: "ПЯТНИЦА"}

        # Собираем списки групп и преподавателей (для селектов)
        all_groups = set()
        all_teachers = set()

        if "ФИО преподавателя" in self.teachers.columns:
            for t in self.teachers["ФИО преподавателя"]:
                if pd.notna(t):
                    all_teachers.add(str(t).strip())

        groups_sorted = sorted(all_groups)
        teachers_sorted = sorted(all_teachers)

        weeks_out = []
        for w in range(1, self.WEEKS + 1):
            if w not in semester_schedule:
                continue

            week_obj = {"week_number": w, "days": []}
            for d in range(1, self.DAYS_PER_WEEK + 1):
                day_obj = {"day_name": days_names.get(d, f"DAY_{d}"), "pairs": []}

                for p in range(1, self.MAX_PAIRS + 1):
                    # shift: 1-я смена (1-3 пары), 2-я смена (4-5 пары) — как у вас в UI
                    shift = 2 if p >= 4 else 1

                    slots = {}
                    # Инициализируем слоты по всем группам (чтобы фронт мог обращаться pairData.slots[group])
                    for g in groups_sorted:
                        slots[g] = None

                    # Заполняем фактические занятия
                    for g, info in semester_schedule[w][d][p].items():
                        if not info:
                            continue
                        slots[g] = {
                            "subject": info.get("subject", ""),
                            "teacher": info.get("teacher", ""),
                            "room": info.get("room", ""),
                            "is_flow": bool(info.get("is_flow", False)),
                            "shift": shift
                        }

                    day_obj["pairs"].append({
                        "pair": p,
                        "slots": slots
                    })

                week_obj["days"].append(day_obj)

            weeks_out.append(week_obj)

        payload = {
            "groups": groups_sorted,
            "teachers": sorted(all_teachers),
            "weeks": weeks_out
        }

        out_path = Path(output_json_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        with out_path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        print(f"JSON успешно сохранён: {out_path.resolve()}")
def build_json_for_one_semester(optimizer, semester_schedule):
    days_names = {1: "ПОНЕДЕЛЬНИК", 2: "ВТОРНИК", 3: "СРЕДА", 4: "ЧЕТВЕРГ", 5: "ПЯТНИЦА"}

    all_groups = set()
    all_teachers = set()

    for w in semester_schedule:
        for d in semester_schedule[w]:
            for p in semester_schedule[w][d]:
                for g, info in semester_schedule[w][d][p].items():
                    all_groups.add(g)
                    if info and info.get("teacher"):
                        all_teachers.add(info["teacher"])

    groups_sorted = sorted(all_groups)
    teachers_sorted = sorted(all_teachers)

    weeks_out = []
    for w in range(1, optimizer.WEEKS + 1):
        if w not in semester_schedule:
            continue

        week_obj = {"week_number": w, "days": []}
        for d in range(1, optimizer.DAYS_PER_WEEK + 1):
            day_obj = {"day_name": days_names.get(d, f"DAY_{d}"), "pairs": []}

            for p in range(1, optimizer.MAX_PAIRS + 1):
                shift = 2 if p >= 4 else 1
                slots = {g: None for g in groups_sorted}

                for g, info in semester_schedule[w][d][p].items():
                    if not info:
                        continue
                    slots[g] = {
                        "subject": info.get("subject", ""),
                        "teacher": info.get("teacher", ""),
                        "room": info.get("room", ""),
                        "is_flow": bool(info.get("is_flow", False)),
                        "shift": shift
                    }

                day_obj["pairs"].append({"pair": p, "slots": slots})

            week_obj["days"].append(day_obj)

        weeks_out.append(week_obj)

    return {"groups": groups_sorted, "teachers": teachers_sorted, "weeks": weeks_out}
def generate_schedule_from_excel(file_path: str):
    df_rup = pd.read_excel(file_path, sheet_name='РУП')
    df_teachers = pd.read_excel(file_path, sheet_name='Нагруженность преподователей')
    df_groups = pd.read_excel(file_path, sheet_name='Группы и направления')
    df_rooms = pd.read_excel(file_path, sheet_name='Аудитории')
    try:
        df_rules = pd.read_excel(file_path, sheet_name='Правила составления')
    except:
        df_rules = pd.DataFrame()

    # гарантируем спортзал
    has_gym = False
    if 'Аудитория' in df_rooms.columns:
        has_gym = any('спорт' in str(r).lower() for r in df_rooms['Аудитория'])
    if not has_gym:
        new_room = pd.DataFrame([{'Аудитория': 'Спорт зал', 'Назначение': 'Физра', 'Вместимость': 100}])
        df_rooms = pd.concat([df_rooms, new_room], ignore_index=True)

    # --- SPLIT BY SEMESTER ---
    loads_by_semester = split_load_by_semester(df_teachers)

    semesters_payload = {}
    warnings_by_semester = {}

    for sem, df_load_sem in loads_by_semester.items():
        optimizer = ScheduleOptimizer(df_rup, df_load_sem, df_groups, df_rooms, df_rules)
        semester_sched, warnings = optimizer.generate_semester()

        # строим JSON одного семестра (логика как в вашем save_semester_to_json)
        sem_payload = build_json_for_one_semester(optimizer, semester_sched)
        semesters_payload[str(sem)] = sem_payload
        warnings_by_semester[str(sem)] = warnings

    final_payload = {"semesters": semesters_payload}

    json_path = os.path.join("vue-project", "public", "schedule_data.json")
    Path(json_path).parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(final_payload, f, ensure_ascii=False, indent=2)

    return {"json_path": json_path, "warnings": warnings_by_semester}
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
@app.post("/process")
async def process(file: UploadFile = File(...)):
    excel_bytes = await file.read()

    tech_report = ai_validate_excel(excel_bytes)
    if tech_report.get("summary", {}).get("errors", 0) > 0:
        return JSONResponse(
            status_code=400,
            content={"ok": False, "stage": "tech_validation_failed", "report": tech_report}
        )

    logic_errors = logic_precheck_full(excel_bytes)
    if logic_errors:
        logic_report = {
            "summary": {"errors": len(logic_errors), "warnings": 0, "notes": 0},
            "errors": logic_errors,
            "warnings": [],
            "notes": [],
            "rules_feedback": {"params": [], "hard": [], "soft": [], "issues": [], "suggestions": []},
        }
        return JSONResponse(
            status_code=400,
            content={"ok": False, "stage": "logic_validation_failed", "report": logic_report}
        )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_xlsx:
        tmp_xlsx.write(excel_bytes)
        tmp_path = tmp_xlsx.name

    try:
        result = generate_schedule_from_excel(tmp_path)

        # ВАЖНО: вернуть JSON в ответ (а не файл)
        json_path = result["json_path"]
        with open(json_path, "r", encoding="utf-8") as f:
            schedule_json = json.load(f)

        return JSONResponse(
            status_code=200,
            content={"ok": True, "stage": "generated", "data": schedule_json, "warnings": result["warnings"]}
        )
    finally:
        try:
            os.remove(tmp_path)
        except:
            pass
# @app.post("/generate")
# async def generate(file: UploadFile = File(...)):
#     excel_bytes = await file.read()

#     # сохраняем загруженный xlsx во временный файл
#     with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
#         tmp.write(excel_bytes)
#         tmp_path = tmp.name

#     try:
#         result = generate_schedule_from_excel(tmp_path)
#         return {"ok": True, "saved_to": result["json_path"], "warnings": result["warnings"]}
#     finally:
#         try:
#             os.remove(tmp_path)
#         except:
#             pass
# ==========================================
# ЗАПУСК
# ==========================================
# if __name__ == "__main__":
#     file_path = 'Книга (3) (1).xlsx' 
    
#     if not os.path.exists(file_path):
#         print(f"ОШИБКА: Файл '{file_path}' не найден!")
#     else:
#         try:
#             df_rup = pd.read_excel(file_path, sheet_name='РУП') 
#             df_teachers = pd.read_excel(file_path, sheet_name='Нагруженность преподователей') 
#             df_groups = pd.read_excel(file_path, sheet_name='Группы и направления') 
#             df_rooms = pd.read_excel(file_path, sheet_name='Аудитории') 
#             try: df_rules = pd.read_excel(file_path, sheet_name='Правила составления')
#             except: df_rules = pd.DataFrame()

#             has_gym = False
#             for r in df_rooms['Аудитория']:
#                 if 'спорт' in str(r).lower(): has_gym = True; break
#             if not has_gym:
#                 new_room = pd.DataFrame([{'Аудитория': 'Спорт зал', 'Назначение': 'Физра', 'Вместимость': 100}])
#                 df_rooms = pd.concat([df_rooms, new_room], ignore_index=True)

#             optimizer = ScheduleOptimizer(df_rup, df_teachers, df_groups, df_rooms, df_rules)
#             semester_sched, errors = optimizer.generate_semester()
            
#             if errors:
#                 print(f"\nЕсть {len(errors)} предупреждений.")
#                 for e in errors: print(f" - {e}")
#             else:
#                 print("\nУСПЕХ: Все недели сгенерированы без ошибок!")

#             optimizer.save_semester_to_excel(semester_sched, "Умное_Расписание_Final.xlsx")

#             # JSON для Vue
#             json_path = os.path.join("vue-project", "public", "schedule_data.json")
#             optimizer.save_semester_to_json(semester_sched, json_path)

#         except Exception as e:
#             import traceback
#             print(f"\nКРИТИЧЕСКАЯ ОШИБКА: {e}")
#             traceback.print_exc()